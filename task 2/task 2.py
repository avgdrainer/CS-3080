import requests, bs4, openpyxl, re
from openpyxl.utils import get_column_letter

class Author:

  def __init__(self, name):
      self.name = name
      self.contributions = {"2021": 0, "2022": 0, "2023": 0}
    
  def tally(self, year):
      self.contributions[year] += 1

  def sum(self):
    return sum(self.contributions.values())



def main():

  authors = {} 
  urls = [r'https://openaccess.thecvf.com/CVPR2021?day=all', r'https://openaccess.thecvf.com/CVPR2022?day=all', r'https://openaccess.thecvf.com/CVPR2023?day=all']

  for url in urls:
    
    year = getYear(url)
    names = getNames(url)
    countNames (names, authors, year)

  topThree = topContributors(authors)
  sendExcel(topThree)

#uses pattern matching to determine what year we are looking at
def getYear(url):

  pattern = re.compile(r'\d{4}')
  year = pattern.search(url)

  return year.group()

#creates a BeautifulSoup object, scans for all contributors, places in list
def getNames(url):

  res = requests.get(url)

  res.raise_for_status()

  bs4_object = bs4.BeautifulSoup(res.text, 'html.parser')
  names = bs4_object.select('input[name="query_author"]')  

  return names

#takes list of names, for each unique name creates an author object, adds to author list
# for each instance of that name in the list, increment Author.contributions, del name
def countNames(names, authors, year):

   for name in names: #for x in range(12):

    value = name.get('value')

    if value in authors:
      authors[value].tally(year)
    else:
      authors[value] = Author(value)
      authors[value].tally(year)

#iterates through authors dictionary values to find top three, return list of corresponding Author objects 
def topContributors(authors):

  
  if len(authors) < 3:
    print("not enough contributors for a top three. exiting program...")
    exit()
  elif len(authors) == 3:
    return authors
  else:
    authorObjects = list(authors.values())
    authorObjects.sort(key=lambda x: x.sum(), reverse=True)
    authorObjects = authorObjects[:3]
    return authorObjects     
      
#prints the top three authors' number of contributions in an Excel spreadsheet
def sendExcel(topThree):
 
  wb = openpyxl.Workbook()
  sheet = wb.active

  sheet['A2'] = "2021"
  sheet['A3'] = "2022"
  sheet['A4'] = "2023"
  sheet['A5'] = "Total"

  for x in range(0, 3):
      
    column = get_column_letter(2 + x)
    sheet[column + '1'] = topThree[x].name

    for y in range(2, 5):
      sheet[column + f'{y}'] = topThree[x].contributions[sheet['A' + f'{y}'].value]
    sheet[column + '5'] = topThree[x].sum()

  wb.save(r'C:\Users\Riley\Documents\GitHub\CS-3080\task 2\top three.xlsx')


main()

